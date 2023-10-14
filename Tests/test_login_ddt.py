import pytest
import time
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium.webdriver.common.by import By
from Testdata.testdata import Testdata
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# Read test data from Excel file
workbook = load_workbook('Testdata/User_LoGIN.xlsx')
sheet = workbook['Worksheet']

# Define logging xl function
def log_login(email, password, status, reason):
    # sourcery skip: hoist-similar-statement-from-if, hoist-statement-from-if
    if 'logs2' not in workbook.sheetnames:
        new_sheet = workbook.create_sheet('logs2')  # Create new worksheet
    new_sheet = workbook['logs2']
    # Write headers to new sheet
    new_sheet.cell(row=1, column=1, value='Email')
    new_sheet.cell(row=1, column=2, value='Password')
    new_sheet.cell(row=1, column=3, value='Status')
    new_sheet.cell(row=1, column=4, value='Reason') 
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for col in range(1, 5):
        cell = new_sheet.cell(row=1, column=col)
        cell.fill = yellow_fill

    row = new_sheet.max_row + 1
    new_sheet.cell(row=row, column=1).value = email
    new_sheet.cell(row=row, column=2).value = password
    new_sheet.cell(row=row, column=3).value = status
    new_sheet.cell(row=row, column=4).value = reason
    if status == 'Success':
        fill = PatternFill(start_color='00FF00',
                           end_color='00FF00', fill_type='solid')
        new_sheet.cell(row=row, column=3).fill = fill
    else:
        fill = PatternFill(start_color='FF0000',
                           end_color='FF0000', fill_type='solid')
        new_sheet.cell(row=row, column=3).fill = fill
    workbook.save('Testdata/User_LoGIN.xlsx')


@pytest.mark.parametrize('email, password', [(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value) for i in range(2, sheet.max_row + 1)])
def test_login(email, password):  # sourcery skip: hoist-similar-statement-from-if, hoist-statement-from-if
    if not hasattr(test_login, 'driver'):
        # using hasattr(),This attribute is used to store the browser instance, and
        # subsequent iterations of the test will reuse the same browser instance
        test_login.driver = webdriver.Chrome()  # Initialize driver

    driver = test_login.driver
    driver.get(Testdata.base_url)

    
    email_input = driver.find_element(By.XPATH, "//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div[2]/input")
    email_input.send_keys(email)
    passwd_input = driver.find_element(By.XPATH, "//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[2]/div/div[2]/input")
    passwd_input.send_keys(password)
    # remember_me=driver.find_element(By.CSS_SELECTOR,"input[type='checkbox']")
    # remember_me.click()
    # time.sleep(2)
    login_button = driver.find_element(By.CSS_SELECTOR, "input[value='Login']")
    login_button.click()
    time.sleep(4)
    
    # driver.get_screenshot_as_file("Screenshots/login1.png")
    alert = driver.switch_to.alert
    alert_text = alert.text
    # time.sleep(4)


    if alert_text =="Loggedin Successfully":
        status = 'Success'
        reason = alert_text#'Success'
        print(reason)
        alert.accept()
        WebDriverWait(driver, 10).until(EC.url_to_be(Testdata.expected_url))
        print(driver.current_url)
        # driver.save_screenshot("Screenshots/success.png")# has scrollbar
        assert driver.current_url == Testdata.expected_url,"Failed"
        

    else:
        print(alert.text)
        status = 'Failure'
        reason = alert.text#'Failure'
        alert.accept()
        # driver.save_screenshot("Screenshots/failure.png")
        assert driver.current_url != Testdata.expected_url,"Passed"
    log_login(email, password, status, reason)



